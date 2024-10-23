import openpyxl
import xlwings as xw
import os
import argparse
import zipfile  # Import for handling zip-related errors

# Set up argument parser to accept paths from the user
parser = argparse.ArgumentParser(description="Process Excel files and export to PDF.")
parser.add_argument('source_folder', type=str, help='Path to the folder containing source Excel files')
parser.add_argument('target_template', type=str, help='Path to the target Excel template file')
parser.add_argument('output_folder', type=str, help='Path to the folder where output PDFs will be saved')

# Parse the arguments
args = parser.parse_args()

# Define paths from user inputs
source_folder = args.source_folder
target_template = args.target_template
output_folder = args.output_folder

# Create the output folder if it doesn't exist
os.makedirs(output_folder, exist_ok=True)

# Change the current working directory to source_folder
os.chdir(source_folder)

# Iterate over all .xlsx files in the source folder
for source_file in os.listdir():
    if source_file.lower().endswith('.xlsx'):
        source_path = os.path.join(source_folder, source_file)
        print(f"Attempting to load: {source_path}")
        
        try:
            # Load the source workbook and the first sheet
            source_wb = openpyxl.load_workbook(source_path, data_only=True)
        except zipfile.BadZipFile:
            print(f"Skipping file {source_file}: it is not a valid .xlsx file.")
            continue

        # Get the value from cell H1 to use as part of the PDF name
        source_ws = source_wb.active  # Get the first (active) sheet from source
        pdf_name_prefix = source_ws['H1'].value
        if not pdf_name_prefix:
            print(f"Skipping file {source_file}: H1 is empty.")
            continue

        # Define the output PDF name
        pdf_output = os.path.join(output_folder, f"{pdf_name_prefix}_corrected.pdf")

        # Load the target template workbook
        target_wb = openpyxl.load_workbook(target_template)
        target_ws = target_wb.active  # Automatically get the first (active) sheet

        # Preserve merged cell ranges from source to target
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))

        # Copy content from source to target, handling merged cells
        for row in source_ws.iter_rows():
            for cell in row:
                # Check if the cell is the top-left cell of a merged range or if it is not in any merged range
                if not any(cell.coordinate in merged_range for merged_range in source_ws.merged_cells.ranges) or \
                   cell.coordinate == next(iter(source_ws.merged_cells.ranges), None).min_row:
                    target_cell = target_ws.cell(row=cell.row, column=cell.column)
                    target_cell.value = cell.value

        # Save the updated target workbook as a temporary file
        temp_target_path = os.path.join(output_folder, f"{pdf_name_prefix}_temp.xlsx")
        target_wb.save(temp_target_path)
        print(f"Data from {source_file} has been copied to the template.")

        # Open the updated target file with xlwings for PDF export
        app = xw.App(visible=False)
        wb = app.books.open(temp_target_path)

        # Reference Sheet2 for export to PDF
        sheet_to_export = wb.sheets[1]  # Adjust index if the sheet is different

        # Export the selected sheet as a PDF
        sheet_to_export.to_pdf(pdf_output)
        print(f"Exported {pdf_output}")

        # Close the workbook and quit the app
        wb.close()
        app.quit()

        # Remove the temporary Excel file
        os.remove(temp_target_path)
        print(f"Removed temporary file: {temp_target_path}")

print("All files have been processed.")
